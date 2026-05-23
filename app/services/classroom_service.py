"""
app/services/classroom_service.py
教室资源 CRUD 业务逻辑。
"""

import csv
import io
from dataclasses import dataclass
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.classroom import Classroom, ClassroomType
from app.schemas.classroom import (
    ClassroomBatchImportResult,
    ClassroomCreate,
    ClassroomImportFailure,
    ClassroomUpdate,
)
from app.schemas.response import BizCode, BizException

MAX_IMPORT_FILE_SIZE = 5 * 1024 * 1024
REQUIRED_IMPORT_COLUMNS = {"code", "name", "campus", "building", "capacity"}


@dataclass(slots=True)
class _ClassroomImportRow:
    row: int
    code: str
    data: dict[str, Any]


async def create_classroom(db: AsyncSession, data: ClassroomCreate) -> Classroom:
    existing = await db.scalar(select(Classroom).where(Classroom.code == data.code))
    if existing:
        raise BizException(
            BizCode.GENERAL_ERROR,
            f"Classroom code '{data.code}' already exists",
        )
    obj = Classroom(**data.model_dump())
    db.add(obj)
    await db.commit()
    await db.refresh(obj)
    return obj


async def get_classroom(db: AsyncSession, classroom_id: int) -> Classroom:
    obj = await db.get(Classroom, classroom_id)
    if not obj:
        raise BizException(
            BizCode.CLASSROOM_NOT_FOUND,
            f"Classroom {classroom_id} not found",
        )
    return obj


async def list_classrooms(db: AsyncSession, skip: int = 0, limit: int = 100) -> list[Classroom]:
    result = await db.execute(select(Classroom).offset(skip).limit(limit))
    return list(result.scalars().all())


async def update_classroom(db: AsyncSession, classroom_id: int, data: ClassroomUpdate) -> Classroom:
    obj = await get_classroom(db, classroom_id)
    for field, value in data.model_dump(exclude_none=True).items():
        setattr(obj, field, value)
    await db.commit()
    await db.refresh(obj)
    return obj


async def delete_classroom(db: AsyncSession, classroom_id: int) -> None:
    obj = await get_classroom(db, classroom_id)
    await db.delete(obj)
    await db.commit()


async def batch_import_classrooms(
    db: AsyncSession,
    filename: str,
    content: bytes,
    overwrite: bool = False,
) -> ClassroomBatchImportResult:
    rows, failed = _parse_import_file(filename, content)
    success = 0

    for row in rows:
        try:
            changed = await _upsert_import_row(db, row, overwrite=overwrite)
            if changed:
                success += 1
        except SQLAlchemyError as exc:
            failed.append(
                ClassroomImportFailure(
                    row=row.row,
                    code=row.code,
                    error=f"数据库写入失败: {exc.__class__.__name__}",
                )
            )

    try:
        await db.commit()
    except SQLAlchemyError as exc:
        await db.rollback()
        raise BizException(
            BizCode.GENERAL_ERROR,
            f"批量导入提交失败: {exc.__class__.__name__}",
        ) from exc

    return ClassroomBatchImportResult(success=success, failed=failed)


async def _upsert_import_row(
    db: AsyncSession,
    row: _ClassroomImportRow,
    overwrite: bool,
) -> bool:
    async with db.begin_nested():
        existing = await db.scalar(select(Classroom).where(Classroom.code == row.code))
        if existing:
            if not overwrite:
                return False
            for field, value in row.data.items():
                if field != "code":
                    setattr(existing, field, value)
            await db.flush()
            return True

        db.add(Classroom(**row.data))
        await db.flush()
        return True


def _parse_import_file(
    filename: str,
    content: bytes,
) -> tuple[list[_ClassroomImportRow], list[ClassroomImportFailure]]:
    suffix = _file_suffix(filename)
    if suffix == ".csv":
        raw_rows = _read_csv_rows(content)
    elif suffix == ".xlsx":
        raw_rows = _read_xlsx_rows(content)
    else:
        raise BizException(
            BizCode.VALIDATION_ERROR,
            "仅支持 .csv 或 .xlsx 格式的教室导入文件",
        )

    rows: list[_ClassroomImportRow] = []
    failed: list[ClassroomImportFailure] = []
    for row_number, raw_row in raw_rows:
        if _is_empty_row(raw_row):
            continue

        code = _cell_to_str(raw_row.get("code")) or None
        try:
            row = _validate_import_row(row_number, raw_row)
            rows.append(row)
        except (ValueError, ValidationError) as exc:
            failed.append(
                ClassroomImportFailure(
                    row=row_number,
                    code=code,
                    error=_format_validation_error(exc),
                )
            )

    return rows, failed


def _read_csv_rows(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise BizException(BizCode.VALIDATION_ERROR, "CSV 文件必须使用 UTF-8 编码") from exc

    try:
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise BizException(BizCode.VALIDATION_ERROR, "导入文件缺少表头")
        headers = [_normalize_header(header) for header in reader.fieldnames]
        _validate_headers(headers)

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, row in enumerate(reader, start=2):
            normalized = {}
            for header, value in row.items():
                if header is None:
                    continue
                normalized_header = _normalize_header(header)
                if normalized_header:
                    normalized[normalized_header] = value
            rows.append((row_number, normalized))
        return rows
    except csv.Error as exc:
        raise BizException(BizCode.VALIDATION_ERROR, f"CSV 文件解析失败: {exc}") from exc


def _read_xlsx_rows(content: bytes) -> list[tuple[int, dict[str, Any]]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise BizException(BizCode.VALIDATION_ERROR, "Excel 文件解析失败") from exc

    try:
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if header_row is None:
            raise BizException(BizCode.VALIDATION_ERROR, "导入文件缺少表头")

        headers = [_normalize_header(header) for header in header_row]
        _validate_headers(headers)

        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(rows_iter, start=2):
            normalized = {}
            for header, value in zip(headers, values, strict=False):
                if header:
                    normalized[header] = value
            rows.append((row_number, normalized))
        return rows
    finally:
        workbook.close()


def _validate_headers(headers: list[str]) -> None:
    header_set = {header for header in headers if header}
    missing = sorted(REQUIRED_IMPORT_COLUMNS - header_set)
    if missing:
        raise BizException(
            BizCode.VALIDATION_ERROR,
            f"导入文件缺少必填列: {', '.join(missing)}",
        )


def _validate_import_row(row_number: int, raw_row: dict[str, Any]) -> _ClassroomImportRow:
    payload = {
        "code": _required_text(raw_row, "code"),
        "name": _required_text(raw_row, "name"),
        "campus": _required_text(raw_row, "campus"),
        "building": _required_text(raw_row, "building"),
        "capacity": _parse_capacity(raw_row.get("capacity")),
        "room_type": _parse_room_type(raw_row.get("room_type")),
        "available_time": _parse_available_time(raw_row.get("available_time")),
    }
    create_data = ClassroomCreate.model_validate(payload)
    data = create_data.model_dump()
    data["is_active"] = _parse_is_active(raw_row.get("is_active"))
    return _ClassroomImportRow(row=row_number, code=data["code"], data=data)


def _required_text(raw_row: dict[str, Any], field: str) -> str:
    value = _cell_to_str(raw_row.get(field))
    if not value:
        raise ValueError(f"{field} 为必填字段")
    return value


def _parse_capacity(value: Any) -> int:
    if value is None or value == "":
        raise ValueError("capacity 为必填字段")
    if isinstance(value, bool):
        raise ValueError("capacity 必须为正整数")
    if isinstance(value, int):
        capacity = value
    elif isinstance(value, float) and value.is_integer():
        capacity = int(value)
    else:
        text = _cell_to_str(value)
        try:
            capacity = int(text)
        except ValueError as exc:
            try:
                numeric = float(text)
            except ValueError:
                raise ValueError("capacity 必须为正整数") from exc
            if not numeric.is_integer():
                raise ValueError("capacity 必须为正整数") from exc
            capacity = int(numeric)

    if capacity <= 0:
        raise ValueError("capacity 必须为正整数")
    return capacity


def _parse_room_type(value: Any) -> ClassroomType:
    text = _cell_to_str(value)
    if not text:
        return ClassroomType.LECTURE
    try:
        return ClassroomType(text.upper())
    except ValueError as exc:
        allowed = ", ".join(item.value for item in ClassroomType)
        raise ValueError(f"room_type 必须是以下值之一: {allowed}") from exc


def _parse_available_time(value: Any) -> list[dict[str, int]]:
    text = _cell_to_str(value)
    if not text:
        return []

    slots: list[dict[str, int]] = []
    for token in text.split(","):
        token = token.strip()
        if not token:
            continue
        parts = token.split("-")
        if len(parts) != 2:
            raise ValueError("available_time 格式应为 1-1,1-2,2-3")
        try:
            day = int(parts[0].strip())
            slot = int(parts[1].strip())
        except ValueError as exc:
            raise ValueError("available_time 中的 day 和 slot 必须为整数") from exc
        if not 1 <= day <= 7:
            raise ValueError("available_time 中的 day 必须在 1-7 之间")
        if not 1 <= slot <= 12:
            raise ValueError("available_time 中的 slot 必须在 1-12 之间")
        slots.append({"day": day, "slot": slot})
    return slots


def _parse_is_active(value: Any) -> bool:
    if value is None or value == "":
        return True
    if isinstance(value, bool):
        return value

    text = _cell_to_str(value).lower()
    if text in {"true", "1", "yes", "y", "是", "启用", "active"}:
        return True
    if text in {"false", "0", "no", "n", "否", "停用", "inactive"}:
        return False
    raise ValueError("is_active 必须是布尔值")


def _cell_to_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return _cell_to_str(value).lower()


def _file_suffix(filename: str) -> str:
    dot_index = filename.rfind(".")
    if dot_index == -1:
        return ""
    return filename[dot_index:].lower()


def _is_empty_row(row: dict[str, Any]) -> bool:
    return all(not _cell_to_str(value) for value in row.values())


def _format_validation_error(exc: ValueError | ValidationError) -> str:
    if isinstance(exc, ValidationError):
        return "; ".join(
            f"{'.'.join(str(part) for part in error['loc'])}: {error['msg']}"
            for error in exc.errors()
        )
    return str(exc)
